import openpyxl

wb=openpyxl.load_workbook("/storage/emulated/0/gg.xlsx")
ws=wb.active

dat = True
jumlah_data=3
siswa=[]

while dat:
	jumlah_data +=1
	for x in range(jumlah_data):
		data=ws.cell(row=jumlah_data,column=1).value	
		if data == None:
			dat = False
	
	
for x in range(3,jumlah_data):
	ket={}
	
	ket["id"]=ws.cell(row=x,column=1).value
	ket["Nama"]=ws.cell(row=x,column=2).value
	ket["Mtk"]=ws.cell(row=x,column=3).value
	ket["Bahasa"]=ws.cell(row=x,column=4).value
	ket["QH"]=ws.cell(row=x,column=5).value
		
	siswa.append(ket)
print(siswa)
	
		



#wb.save('/storage/emulated/0/gg.xlsx')
print(jumlah_data)