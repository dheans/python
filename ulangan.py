# project baru
# list kunci jawaban
import openpyxl as xl

# create excel file and header of
wb = xl.Workbook()
ws = wb.active
ro = ws.max_row


# memeriksa jawaban siswa
jumlah_siswa = int(input("Masukkan jumlah siswa : "))

key = list(input(
    "Insert answer key :\n separated with space!\nEx : A B C B\n\t").split(" "))

print("Insert students answer with name at the first\n\tEx: Ilma A B C D\n", "*_"*20)

data = []
jumlah_soal = len(key)

while jumlah_siswa > 0:
    dt = {}
    jumlah_siswa -= 1
    jumlah_benar = 0
    answer = input("Insert answer :\n\t").split(" ")

    for x in range(len(key)):
        if key[x] == answer[x+1]:
            jumlah_benar += 1

        dt["Nama"] = answer[0]
        dt["Benar"] = jumlah_benar
        dt["Salah"] = jumlah_soal - jumlah_benar
        dt["Nilai"] = int(jumlah_benar * (100/len(key)))

    data.append(dt)


for x in data:
    print()
    print("Nama -->\t%s\nBenar -->\t%d\nNilai -->\t%.2f" %
          (x["Nama"], x["Benar"], x["Nilai"]))

header = list(dt.keys())
for i in range(len(header)):
    ws.cell(row=1, column=i+1, value=header[i])
for i in data:
    ro += 1
    ws.cell(row=ro, column=1, value=i["Nama"])
    ws.cell(row=ro, column=2, value=i["Benar"])
    ws.cell(row=ro, column=3, value=i["Salah"])
    ws.cell(row=ro, column=4, value=i["Nilai"])
wb.save("/storage/emulated/0/nilai.xlsx")
